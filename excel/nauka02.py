import openpyxl
import openpyxl.styles

wb = openpyxl.Workbook()

ws = wb.active

ws['B5'] = 'Szkolenie Python w ALX'

ft = openpyxl.styles.Font(name="Tahoma", color="ee8800",bold=True)

ws['B5'].font = ft

wb.save("nauka02.xlsx")
